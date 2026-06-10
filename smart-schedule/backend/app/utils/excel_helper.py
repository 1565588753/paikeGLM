"""
Excel 导入导出工具模块
- 使用 openpyxl 处理 Excel 文件
- 支持用户、班级、任课安排、人事表等的导入导出
"""
import io
import os
from datetime import datetime
from typing import List, Dict, Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from app.config import settings


# ==================== 样式定义 ====================

# 标题样式
HEADER_FONT = Font(name="微软雅黑", size=12, bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 普通单元格样式
CELL_FONT = Font(name="微软雅黑", size=10)
CELL_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 边框
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_workbook_with_headers(headers: List[str], sheet_name: str = "Sheet1") -> Workbook:
    """
    创建带有表头的工作簿
    - headers: 表头列表
    - sheet_name: 工作表名称
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # 设置列宽
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    return wb


def write_rows_to_worksheet(ws, rows: List[List[Any]], start_row: int = 2):
    """
    向工作表写入数据行
    - ws: 工作表对象
    - rows: 数据行列表
    - start_row: 起始行号
    """
    for row_idx, row_data in enumerate(rows, start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER


def workbook_to_bytes(wb: Workbook) -> bytes:
    """将工作簿转换为字节流"""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def read_excel_file(file_bytes: bytes, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    读取 Excel 文件内容
    - file_bytes: 文件字节流
    - sheet_name: 工作表名称，默认读取第一个
    - 返回字典列表，键为表头名称
    """
    wb = load_workbook(io.BytesIO(file_bytes))
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    # 读取表头
    headers = []
    for cell in ws[1]:
        if cell.value is not None:
            headers.append(str(cell.value).strip())
        else:
            headers.append(f"column_{cell.column}")

    # 读取数据行
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for idx, value in enumerate(row):
            if idx < len(headers):
                row_dict[headers[idx]] = value
        rows.append(row_dict)

    return rows


# ==================== 课表导出专用 ====================

def export_timetable_to_excel(
    entries: List[Dict[str, Any]],
    days_per_week: int = 5,
    max_period: int = 8,
    title: str = "课表",
) -> bytes:
    """
    导出课表到 Excel
    - entries: 课表条目列表，每个条目包含 day_of_week, period_number, subject_name, teacher_name, classroom_name 等
    - days_per_week: 每周天数
    - max_period: 最大节数
    - title: 标题
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # 星期标题
    day_names = ["", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    # 写入标题行
    ws.cell(row=1, column=1, value="节次").font = HEADER_FONT_WHITE
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).alignment = HEADER_ALIGNMENT
    ws.cell(row=1, column=1).border = THIN_BORDER

    for day in range(1, days_per_week + 1):
        cell = ws.cell(row=1, column=day + 1, value=day_names[day])
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # 按天和节次组织数据
    entry_map: Dict[tuple, List[Dict]] = {}
    for entry in entries:
        key = (entry.get("day_of_week", 0), entry.get("period_number", 0))
        if key not in entry_map:
            entry_map[key] = []
        entry_map[key].append(entry)

    # 写入课表内容
    for period in range(1, max_period + 1):
        # 节次列
        cell = ws.cell(row=period + 1, column=1, value=f"第{period}节")
        cell.font = CELL_FONT
        cell.alignment = CELL_ALIGNMENT
        cell.border = THIN_BORDER

        for day in range(1, days_per_week + 1):
            entries_at_slot = entry_map.get((day, period), [])
            if entries_at_slot:
                # 组合显示：科目/教师/教室
                parts = []
                for e in entries_at_slot:
                    part = e.get("subject_short_name") or e.get("subject_name", "")
                    if e.get("teacher_name"):
                        part += f"\n{e['teacher_name']}"
                    if e.get("classroom_name"):
                        part += f"\n{e['classroom_name']}"
                    parts.append(part)
                value = "\n---\n".join(parts)
            else:
                value = ""

            cell = ws.cell(row=period + 1, column=day + 1, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

    # 设置列宽
    ws.column_dimensions["A"].width = 10
    for day in range(1, days_per_week + 1):
        ws.column_dimensions[get_column_letter(day + 1)].width = 20

    # 设置行高
    for row in range(2, max_period + 2):
        ws.row_dimensions[row].height = 50

    return workbook_to_bytes(wb)


# ==================== 人事表导出专用 ====================

def export_staff_to_excel(
    staff_data: List[Dict[str, Any]],
    class_names: List[str],
    subject_names: List[str],
    title: str = "人事表",
) -> bytes:
    """
    导出人事表到 Excel
    - staff_data: 人事数据列表
    - class_names: 班级名称列表（列头）
    - subject_names: 科目名称列表（行头）
    - title: 标题
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # 写入表头：科目/班级
    ws.cell(row=1, column=1, value="科目").font = HEADER_FONT_WHITE
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).alignment = HEADER_ALIGNMENT
    ws.cell(row=1, column=1).border = THIN_BORDER

    for col_idx, class_name in enumerate(class_names, 2):
        cell = ws.cell(row=1, column=col_idx, value=class_name)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # 构建查找映射: (subject_name, class_name) -> teacher_name
    lookup: Dict[tuple, str] = {}
    for item in staff_data:
        key = (item.get("subject_name", ""), item.get("class_name", ""))
        lookup[key] = item.get("teacher_name", "")

    # 写入数据
    for row_idx, subject_name in enumerate(subject_names, 2):
        cell = ws.cell(row=row_idx, column=1, value=subject_name)
        cell.font = CELL_FONT
        cell.alignment = CELL_ALIGNMENT
        cell.border = THIN_BORDER

        for col_idx, class_name in enumerate(class_names, 2):
            value = lookup.get((subject_name, class_name), "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

    # 设置列宽
    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, len(class_names) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    return workbook_to_bytes(wb)


def ensure_backup_dir() -> str:
    """确保备份目录存在"""
    backup_dir = settings.BACKUP_DIR
    os.makedirs(backup_dir, exist_ok=True)
    return backup_dir


def generate_backup_filename(prefix: str = "backup") -> str:
    """生成备份文件名"""
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{now}.sql"
